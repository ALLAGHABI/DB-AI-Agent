"""تصدير التقرير: Excel منسق وPDF (إن توفر محرك)."""
import io

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="059669")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws):
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for i, col in enumerate(ws.columns, 1):
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 4, 12), 50)


def _sheet_name(name: str, used: set) -> str:
    """أسماء أوراق Excel محدودة بـ31 حرفاً ويجب أن تكون فريدة."""
    base = str(name)[:28] or "sheet"
    candidate, i = base, 1
    while candidate in used:
        candidate = f"{base[:26]}_{i}"
        i += 1
    used.add(candidate)
    return candidate


def _write_business(writer, name: str, business: dict, ar: bool, used: set) -> None:
    """ورقة أرقام الأعمال لكل مصدر — اتجاه وتوزيعات جاهزة لإعادة الاستخدام."""
    rows: list[tuple] = []
    for k in business.get("kpis", []):
        rows.append((k["key"], k.get("column") or "", k["value"], ""))
    tr = business.get("trend")
    if tr:
        for label, value in zip(tr["labels"], tr["values"]):
            rows.append(("trend", tr.get("measure") or "", value, label))
    for bd in business.get("breakdowns", []):
        for label, value, pct in zip(bd["labels"], bd["values"], bd["values_pct"]):
            rows.append((bd["column"], bd.get("measure") or "", value,
                         f"{label} ({pct}%)" if pct is not None else label))
    if not rows:
        return
    cols = (["المؤشر", "العمود", "القيمة", "التصنيف"] if ar
            else ["Metric", "Column", "Value", "Label"])
    pd.DataFrame(rows, columns=cols).to_excel(
        writer, sheet_name=_sheet_name(name, used), index=False)


def _narrative_rows(insights: dict, sections, ar: bool) -> list[tuple]:
    rows: list[tuple] = []
    if "summary" in sections and insights.get("summary"):
        rows.append(("الملخص" if ar else "Summary", insights["summary"]))
    if "findings" in sections:
        for i, f in enumerate(insights.get("findings", []), 1):
            rows.append((f"{'نتيجة' if ar else 'Finding'} {i}", f))
    if "recommendations" in sections:
        for i, r in enumerate(insights.get("recommendations", []), 1):
            rows.append((f"{'توصية' if ar else 'Recommendation'} {i}", r))
    return rows


def to_xlsx(profile: dict, insights: dict, language: str,
            sections=("summary", "findings", "recommendations", "appendix")) -> bytes:
    """ورقة أرقام + سرد. لا ورقة «عينة» — تسرّب بيانات الأفراد في ملف يُتداول."""
    ar = language == "ar"
    if profile.get("kind") == "multi":
        return _to_xlsx_multi(profile, insights, ar, sections)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        summary_rows = [
            ("الصفوف" if ar else "Rows", profile["overview"]["rows"]),
            ("الأعمدة" if ar else "Columns", profile["overview"]["cols"]),
            ("قيم مفقودة %" if ar else "Missing %", profile["overview"]["missing_pct"]),
            ("صفوف مكررة" if ar else "Duplicates", profile["overview"]["duplicate_rows"]),
            ("", ""),
        ] + _narrative_rows(insights, sections, ar)
        pd.DataFrame(summary_rows, columns=["البند" if ar else "Item",
                                            "القيمة" if ar else "Value"]) \
            .to_excel(writer, sheet_name="ملخص" if ar else "Summary", index=False)

        used = {"ملخص" if ar else "Summary"}
        for name, b in (profile.get("business") or {}).items():
            _write_business(writer, name, b, ar, used)

        if "appendix" in sections:
            pd.DataFrame(profile["columns"]).drop(columns=["top_values"], errors="ignore") \
                .to_excel(writer, sheet_name="الأعمدة" if ar else "Columns", index=False)

        for ws in writer.book.worksheets:
            _style_header(ws)
            if ar:
                ws.sheet_view.rightToLeft = True
    return buf.getvalue()


def _to_xlsx_multi(profile: dict, insights: dict, ar: bool, sections) -> bytes:
    buf = io.BytesIO()
    used: set = set()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        rows = [
            ("عدد الجداول" if ar else "Tables", profile["overview"]["tables"]),
            ("إجمالي الصفوف" if ar else "Total rows", profile["overview"]["rows"]),
            ("إجمالي الأعمدة" if ar else "Total columns", profile["overview"]["cols"]),
            ("العلاقات" if ar else "Relationships", profile["overview"]["relationships"]),
            ("", ""),
        ] + _narrative_rows(insights, sections, ar)
        pd.DataFrame(rows, columns=["البند" if ar else "Item", "القيمة" if ar else "Value"]) \
            .to_excel(writer, sheet_name=_sheet_name("ملخص" if ar else "Summary", used),
                      index=False)

        for name, b in (profile.get("business") or {}).items():
            _write_business(writer, name, b, ar, used)

        if "appendix" in sections:
            for ds in profile["datasets"]:
                df = pd.DataFrame(ds["profile"]["columns"]).drop(
                    columns=["top_values"], errors="ignore")
                df.to_excel(writer, sheet_name=_sheet_name(f"{ds['name']}-cols", used),
                            index=False)

        for ws in writer.book.worksheets:
            _style_header(ws)
            if ar:
                ws.sheet_view.rightToLeft = True
    return buf.getvalue()


def pdf_engine() -> str | None:
    """يرجع اسم محرك PDF المتاح أو None."""
    try:
        from playwright.sync_api import sync_playwright  # noqa: F401
        return "playwright"
    except ImportError:
        pass
    try:
        import weasyprint  # noqa: F401
        return "weasyprint"
    except ImportError:
        return None


def to_pdf(html: str) -> bytes | None:
    engine = pdf_engine()
    if engine == "playwright":
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()
            page.set_content(html, wait_until="networkidle")
            pdf = page.pdf(format="A4", print_background=True,
                           margin={"top": "15mm", "bottom": "15mm",
                                   "left": "12mm", "right": "12mm"})
            browser.close()
            return pdf
    if engine == "weasyprint":
        import weasyprint
        return weasyprint.HTML(string=html).write_pdf()
    return None
