import time

import pytest
import respx
from fastapi.testclient import TestClient
from httpx import Response

import app.state as state_mod
from app.main import app
from app.state import AppState

CSV = "city,amount\nالرياض,100\nجدة,200\nالرياض,150\n".encode()

FAKE_INSIGHTS = """## SUMMARY
البيانات تغطي 3 صفوف.
## FINDINGS
- الرياض الأكثر تكراراً
## RECOMMENDATIONS
- تابع جدة
"""


@pytest.fixture
def client(tmp_path, monkeypatch):
    from app.config import settings
    monkeypatch.setattr(settings, "data_dir", str(tmp_path))
    fresh = AppState()
    monkeypatch.setattr(state_mod, "state", fresh)
    import app.api.reports_routes as rr
    import app.api.routes as routes_mod
    monkeypatch.setattr(routes_mod, "state", fresh)
    monkeypatch.setattr(rr, "state", fresh)
    return TestClient(app)


def _generate(client, **body):
    """يبدأ التوليد وينتظر انتهاء المهمة الخلفية."""
    r = client.post("/api/reports/generate", json=body)
    if r.status_code != 200:
        return r, None
    job_id = r.json()["job_id"]
    for _ in range(300):                      # حتى 30 ثانية
        status = client.get(f"/api/reports/jobs/{job_id}").json()
        if status["status"] != "running":
            return r, status
        time.sleep(0.1)
    raise AssertionError("job did not finish")


def _analyze(client):
    r = client.post("/api/reports/analyze",
                    files={"file": ("sales.csv", CSV, "text/csv")})
    assert r.status_code == 200
    return r.json()


def test_analyze_returns_profile_and_token(client):
    body = _analyze(client)
    assert body["token"]
    assert body["profile"]["overview"]["rows"] == 3


def test_analyze_rejects_garbage(client):
    r = client.post("/api/reports/analyze",
                    files={"file": ("x.exe", b"\x00\x01", "application/x")})
    assert r.status_code == 400


@respx.mock
def test_generate_and_archive_lifecycle(client):
    respx.get("http://localhost:11434/api/tags").mock(
        return_value=Response(200, json={"models": [{"name": "gemma3:4b"}]}))
    respx.post("http://localhost:11434/api/chat").mock(
        return_value=Response(200, json={"message": {"content": FAKE_INSIGHTS}}))

    token = _analyze(client)["token"]
    _, status = _generate(client, token=token, title="تقرير المبيعات",
                          template="detailed", language="ar",
                          provider="ollama", model="gemma3:4b")
    assert status["status"] == "done", status
    meta = status["report"]
    assert meta["is_local"] is True and meta["rows"] == 3

    # القائمة
    lst = client.get("/api/reports").json()
    assert lst[0]["id"] == meta["id"]

    # HTML
    html = client.get(f"/api/reports/{meta['id']}/html")
    assert html.status_code == 200
    assert "تقرير المبيعات" in html.text
    assert "الرياض الأكثر تكراراً" in html.text

    # Excel
    xlsx = client.get(f"/api/reports/{meta['id']}/xlsx")
    assert xlsx.content[:2] == b"PK"

    # حذف
    assert client.delete(f"/api/reports/{meta['id']}").json()["success"] is True
    assert client.get("/api/reports").json() == []


def test_generate_with_bad_token_404(client):
    r = client.post("/api/reports/generate", json={
        "token": "nope", "title": "x", "provider": "ollama", "model": "m"})
    assert r.status_code == 404


@respx.mock
def test_analyze_table_directly(client, tmp_path):
    """تقرير من جدول متصل — بلا تصدير/رفع."""
    import sqlite3
    db = tmp_path / "shop.db"
    con = sqlite3.connect(db)
    con.executescript(
        "CREATE TABLE sales (id INTEGER PRIMARY KEY, city TEXT, amount REAL);"
        "INSERT INTO sales (city, amount) VALUES ('الرياض', 100), ('جدة', 250), ('الرياض', 75);")
    con.commit(); con.close()
    client.post("/api/db/connect", json={"url": f"sqlite:///{db}"})

    r = client.post("/api/reports/analyze-table", json={"tables": ["sales"]})
    assert r.status_code == 200
    body = r.json()
    assert body["token"]
    assert body["profile"]["overview"]["rows"] == 3

    # جدول غير موجود يعطي رمز خطأ
    r = client.post("/api/reports/analyze-table", json={"tables": ["nope"]})
    assert r.status_code == 404 and r.json()["detail"]["code"] == "tableMissing"


def test_analyze_table_requires_connection(client):
    r = client.post("/api/reports/analyze-table", json={"tables": ["x"]})
    assert r.status_code == 400 and r.json()["detail"]["code"] == "notConnected"


@respx.mock
def test_analyze_whole_database_multi_report(client, tmp_path):
    """قاعدة كاملة: عدة جداول + علاقات في تقرير واحد."""
    import sqlite3
    db = tmp_path / "shop.db"
    con = sqlite3.connect(db)
    con.executescript("""
        CREATE TABLE cities (id INTEGER PRIMARY KEY, name TEXT);
        CREATE TABLE sales (id INTEGER PRIMARY KEY, amount REAL,
                            city_id INTEGER REFERENCES cities(id));
        INSERT INTO cities (name) VALUES ('الرياض'), ('جدة');
        INSERT INTO sales (amount, city_id) VALUES (100, 1), (250, 2), (75, 1);
    """)
    con.commit(); con.close()
    client.post("/api/db/connect", json={"url": f"sqlite:///{db}"})

    # جداول فارغة القائمة = القاعدة كاملة
    r = client.post("/api/reports/analyze-table", json={"tables": []})
    assert r.status_code == 200
    profile = r.json()["profile"]
    assert profile["kind"] == "multi"
    assert profile["overview"]["tables"] == 2
    assert profile["overview"]["rows"] == 5          # 2 مدن + 3 مبيعات
    assert profile["overview"]["relationships"] == 1
    names = {d["name"] for d in profile["datasets"]}
    assert names == {"cities", "sales"}

    # توليد تقرير مركّب فعلي
    respx.get("http://localhost:11434/api/tags").mock(
        return_value=Response(200, json={"models": [{"name": "gemma3:4b"}]}))
    respx.post("http://localhost:11434/api/chat").mock(
        return_value=Response(200, json={"message": {"content": FAKE_INSIGHTS}}))
    _, status = _generate(client, token=r.json()["token"], title="تقرير القاعدة",
                          template="detailed", language="ar",
                          provider="ollama", model="gemma3:4b")
    assert status["status"] == "done", status
    meta = status["report"]

    html = client.get(f"/api/reports/{meta['id']}/html").text
    assert "cities" in html and "sales" in html          # قسم لكل جدول
    assert "العلاقات" in html                             # قسم العلاقات
    assert "chart-sales-1" in html or "chart-cities-1" in html

    import io

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(client.get(f"/api/reports/{meta['id']}/xlsx").content))
    assert "cities" in wb.sheetnames and "sales" in wb.sheetnames


def test_analyze_subset_of_tables(client, tmp_path):
    import sqlite3
    db = tmp_path / "s.db"
    con = sqlite3.connect(db)
    con.executescript(
        "CREATE TABLE a (id INTEGER PRIMARY KEY, v REAL);"
        "CREATE TABLE b (id INTEGER PRIMARY KEY, v REAL);"
        "CREATE TABLE c (id INTEGER PRIMARY KEY, v REAL);"
        "INSERT INTO a (v) VALUES (1),(2); INSERT INTO b (v) VALUES (3);"
        "INSERT INTO c (v) VALUES (4);")
    con.commit(); con.close()
    client.post("/api/db/connect", json={"url": f"sqlite:///{db}"})

    profile = client.post("/api/reports/analyze-table",
                          json={"tables": ["a", "b"]}).json()["profile"]
    assert profile["kind"] == "multi"
    assert {d["name"] for d in profile["datasets"]} == {"a", "b"}

    # جدول واحد يبقى تقريراً مفرداً كامل التفاصيل
    single = client.post("/api/reports/analyze-table",
                         json={"tables": ["a"]}).json()["profile"]
    assert single["kind"] == "single" and single["overview"]["rows"] == 2


@respx.mock
def test_generation_runs_as_a_background_job(client, tmp_path):
    """التوليد يعيد معرّف مهمة فوراً — لا ينتظر النموذج ولا PDF."""
    respx.get("http://localhost:11434/api/tags").mock(
        return_value=Response(200, json={"models": [{"name": "gemma3:4b"}]}))
    respx.post("http://localhost:11434/api/chat").mock(
        return_value=Response(200, json={"message": {"content": FAKE_INSIGHTS}}))

    token = _analyze(client)["token"]
    r = client.post("/api/reports/generate", json={
        "token": token, "title": "خلفية", "provider": "ollama", "model": "gemma3:4b"})
    assert r.status_code == 200
    assert r.json()["status"] == "running" and r.json()["job_id"]

    _, status = _generate(client, token=_analyze(client)["token"], title="خلفية2",
                          provider="ollama", model="gemma3:4b")
    assert status["status"] == "done" and status["report"]["id"]


def test_job_failure_is_reported_with_a_code(client):
    """فشل النموذج يصل كرمز خطأ مترجَم لا كنص خام."""
    token = _analyze(client)["token"]
    _, status = _generate(client, token=token, title="فشل",
                          provider="ollama", model="missing-model")
    assert status["status"] == "failed"
    assert status["error"]["code"] == "insightsFailed"


def test_unknown_job_is_404(client):
    assert client.get("/api/reports/jobs/nope").status_code == 404


def test_analyze_returns_descriptive_labels(client, tmp_path):
    """التقرير يتحدث لغة الأعمال: total_amount → قيمة المبيعات."""
    import sqlite3
    db = tmp_path / "shop.db"
    con = sqlite3.connect(db)
    con.executescript(
        "CREATE TABLE orders (id INTEGER PRIMARY KEY, city TEXT,"
        " order_date TEXT, total_amount REAL);"
        "INSERT INTO orders (city, order_date, total_amount)"
        " VALUES ('الرياض','2025-01-01',100),('جدة','2025-02-01',250);")
    con.commit(); con.close()
    client.post("/api/db/connect", json={"url": f"sqlite:///{db}"})

    labels = client.post("/api/reports/analyze-table",
                         json={"tables": ["orders"], "language": "ar"}).json()["labels"]
    assert labels["total_amount"] == "قيمة المبيعات"
    assert labels["city"] == "المدينة"
    assert labels["order_date"] == "تاريخ الطلب"
    assert labels["orders"] == "الطلبات"

    en = client.post("/api/reports/analyze-table",
                     json={"tables": ["orders"], "language": "en"}).json()["labels"]
    assert en["total_amount"] == "Sales value"


@respx.mock
def test_report_uses_labels_not_raw_column_names(client, tmp_path):
    import sqlite3
    db = tmp_path / "shop2.db"
    con = sqlite3.connect(db)
    con.executescript(
        "CREATE TABLE orders (id INTEGER PRIMARY KEY, city TEXT, total_amount REAL);"
        "INSERT INTO orders (city, total_amount) VALUES ('الرياض',100),('جدة',250);")
    con.commit(); con.close()
    client.post("/api/db/connect", json={"url": f"sqlite:///{db}"})
    token = client.post("/api/reports/analyze-table",
                        json={"tables": ["orders"]}).json()["token"]

    respx.get("http://localhost:11434/api/tags").mock(
        return_value=Response(200, json={"models": [{"name": "gemma3:4b"}]}))
    respx.post("http://localhost:11434/api/chat").mock(
        return_value=Response(200, json={"message": {"content": FAKE_INSIGHTS}}))

    _, status = _generate(client, token=token, title="ت", template="executive",
                          language="ar", provider="ollama", model="gemma3:4b",
                          labels={"city": "الفرع"})
    assert status["status"] == "done", status
    html = client.get(f"/api/reports/{status['report']['id']}/html").text
    assert "قيمة المبيعات" in html          # تسمية تلقائية
    assert "الفرع" in html                   # تسمية كتبها المستخدم
    assert "total_amount" not in html.split("<script")[0]   # لا أسماء خام في المتن


def _multi_sheet_excel() -> bytes:
    """ملف Excel واقعي: ورقتان، وعنوان تزييني قبل صف العناوين."""
    import io

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "المبيعات"
    ws.append(["تقرير الشركة الشهري"])          # صف تزييني
    ws.append([])                                # صف فارغ
    ws.append(["المدينة", "total_amount"])       # صف العناوين الحقيقي
    for city, amount in [("الرياض", 500), ("جدة", 300), ("الرياض", 700)]:
        ws.append([city, amount])

    ws2 = wb.create_sheet("الموظفون")
    ws2.append(["department", "salary"])
    for dept, salary in [("IT", 9000), ("HR", 6000), ("IT", 11000)]:
        ws2.append([dept, salary])

    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def test_all_excel_sheets_are_analyzed(client):
    """ملف بعدة أوراق كان يُقرأ منه الأول فقط."""
    r = client.post("/api/reports/analyze",
                    files={"file": ("company.xlsx", _multi_sheet_excel(), "application/x")},
                    data={"language": "ar"})
    assert r.status_code == 200, r.text
    body = r.json()
    assert set(body["sheets"]) == {"المبيعات", "الموظفون"}
    assert body["profile"]["kind"] == "multi"
    assert body["profile"]["overview"]["tables"] == 2
    assert body["profile"]["overview"]["rows"] == 6          # 3 + 3


def test_decorative_rows_do_not_become_headers(client):
    """صف العنوان التزييني كان يصير أسماء أعمدة مشوّهة."""
    body = client.post("/api/reports/analyze",
                       files={"file": ("company.xlsx", _multi_sheet_excel(), "application/x")},
                       data={"language": "ar"}).json()
    sales = next(d for d in body["profile"]["datasets"] if d["name"] == "المبيعات")
    columns = {c["name"] for c in sales["profile"]["columns"]}
    assert columns == {"المدينة", "total_amount"}
    assert body["labels"]["total_amount"] == "قيمة المبيعات"


@respx.mock
def test_report_stays_arabic_when_model_answers_in_english(client):
    """النموذج العنيد لم يعد يُخرج تقريراً إنجليزياً على مستخدم عربي."""
    english = "## SUMMARY\nThis report analyzes employee working hours and costs.\n"
    respx.get("http://localhost:11434/api/tags").mock(
        return_value=Response(200, json={"models": [{"name": "gemma3:4b"}]}))
    respx.post("http://localhost:11434/api/chat").mock(
        return_value=Response(200, json={"message": {"content": english}}))

    token = client.post("/api/reports/analyze",
                        files={"file": ("company.xlsx", _multi_sheet_excel(), "application/x")},
                        data={"language": "ar"}).json()["token"]
    _, status = _generate(client, token=token, title="تقرير", template="executive",
                          language="ar", provider="ollama", model="gemma3:4b")
    assert status["status"] == "done", status
    meta = status["report"]
    assert meta["language_ok"] is True and meta["used_fallback"] is True

    html = client.get(f"/api/reports/{meta['id']}/html").text
    assert "This report analyzes" not in html
    assert "يغطي هذا التقرير" in html
